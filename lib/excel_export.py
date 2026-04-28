"""
Three-sheet Excel writer for the weekly archive.

For each report (Meta / SEO / AI Readiness):
  Sheet 1 — Summary    (top-line metrics for the dashboard)
  Sheet 2 — Raw        (one row per record)
  Sheet 3 — Delta      (vs previous week's file if present, else a notice)

Public API:
  build_meta_report(week_label) -> Path
  build_seo_report(week_label)  -> Path
  build_ai_readiness_report(week_label) -> Path
  build_all_reports(week_label) -> dict[str, Path]
"""

from __future__ import annotations

import json
import re
from datetime import date, datetime
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from lib.supabase_client import current_iso_week
from lib.synthesis import SEO_CLUSTERS, _load_json

ROOT = Path(__file__).parent.parent
DATA_DIR = ROOT / "data"
REPORTS_DIR = ROOT / "reports"
REPORTS_DIR.mkdir(exist_ok=True)

ADS_FILE = DATA_DIR / "ads_scraped_latest.json"
SEO_FILE = DATA_DIR / "seo_raw_latest.json"
AI_FILE = DATA_DIR / "ai_readiness_latest.json"

HEADER_FILL = PatternFill("solid", fgColor="1C1E26")
HEADER_FONT = Font(name="Calibri", bold=True, color="F0F0F0", size=11)
TITLE_FONT = Font(name="Calibri", bold=True, color="4F8EF7", size=14)


# ── Generic helpers ───────────────────────────────────────────────────────────
def _autosize(ws) -> None:
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[col_letter].width = min(max(12, max_len + 2), 60)


def _write_header(ws, row: int, headers: list[str]) -> None:
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=i, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")


def _write_title(ws, title: str) -> None:
    ws.cell(row=1, column=1, value=title).font = TITLE_FONT
    ws.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M UTC')}")


def _previous_report(prefix: str, current: Path) -> Optional[Path]:
    """Find the most recent report whose name starts with prefix, excluding current."""
    candidates = sorted(REPORTS_DIR.glob(f"{prefix}*.xlsx"), reverse=True)
    for c in candidates:
        if c.resolve() != current.resolve():
            return c
    return None


def _read_xlsx_sheet(path: Path, sheet_name: str) -> list[list]:
    """Read all rows from a worksheet as plain lists."""
    if not path.exists():
        return []
    wb = openpyxl.load_workbook(path, read_only=True)
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    return [list(row) for row in ws.iter_rows(values_only=True) if any(c is not None for c in row)]


# ── META REPORT ───────────────────────────────────────────────────────────────
def build_meta_report(week_label: Optional[str] = None) -> Path:
    week_label = week_label or current_iso_week()
    out = REPORTS_DIR / f"meta_intel_{week_label}.xlsx"
    data = _load_json(ADS_FILE) or {}

    wb = openpyxl.Workbook()

    # Sheet 1 — Summary
    summary = wb.active
    summary.title = "Summary"
    _write_title(summary, f"Listn Meta Intel — {week_label}")
    competitors = data.get("competitors", {})
    total_ads = data.get("total_ads", 0) or sum(len(v) for v in competitors.values())
    active = sum(1 for ads in competitors.values() for a in ads if not a.get("stop_date"))
    summary.cell(row=4, column=1, value="Total ads tracked");      summary.cell(row=4, column=2, value=total_ads)
    summary.cell(row=5, column=1, value="Active ads");             summary.cell(row=5, column=2, value=active)
    summary.cell(row=6, column=1, value="Competitors tracked");    summary.cell(row=6, column=2, value=len(competitors))
    summary.cell(row=7, column=1, value="Data fetched");           summary.cell(row=7, column=2, value=str(data.get("fetched_date", "")))
    summary.cell(row=9, column=1, value="Per-competitor counts").font = Font(bold=True)
    _write_header(summary, 10, ["Competitor", "Ads"])
    for i, (comp, ads) in enumerate(sorted(competitors.items(), key=lambda x: -len(x[1])), start=11):
        summary.cell(row=i, column=1, value=comp)
        summary.cell(row=i, column=2, value=len(ads))
    _autosize(summary)

    # Sheet 2 — Raw
    raw = wb.create_sheet("Raw")
    headers = ["competitor", "ad_id", "page_name", "ad_copy", "headline",
               "cta", "start_date", "stop_date", "days_running", "is_active"]
    _write_header(raw, 1, headers)
    row = 2
    for comp_ads in competitors.values():
        for a in comp_ads:
            raw.cell(row=row, column=1, value=a.get("competitor"))
            raw.cell(row=row, column=2, value=a.get("ad_id"))
            raw.cell(row=row, column=3, value=a.get("page_name"))
            raw.cell(row=row, column=4, value=(a.get("ad_copy") or "")[:1000])
            raw.cell(row=row, column=5, value=a.get("headline"))
            raw.cell(row=row, column=6, value=a.get("cta"))
            raw.cell(row=row, column=7, value=a.get("start_date"))
            raw.cell(row=row, column=8, value=a.get("stop_date"))
            raw.cell(row=row, column=9, value=a.get("days_running"))
            raw.cell(row=row, column=10, value="No" if a.get("stop_date") else "Yes")
            row += 1
    _autosize(raw)
    raw.column_dimensions["D"].width = 80

    # Sheet 3 — Delta
    delta = wb.create_sheet("Delta")
    _write_title(delta, "Week-over-week delta")
    prev = _previous_report("meta_intel_", out)
    if prev is None:
        delta.cell(row=4, column=1, value="No previous week's report found — delta will appear next week.")
    else:
        prev_rows = _read_xlsx_sheet(prev, "Raw")
        prev_ids = {r[1] for r in prev_rows[1:] if r and r[1]}
        curr_ids = {a.get("ad_id") for ads in competitors.values() for a in ads if a.get("ad_id")}
        new_ids = curr_ids - prev_ids
        gone_ids = prev_ids - curr_ids
        delta.cell(row=4, column=1, value=f"Previous week: {prev.name}")
        delta.cell(row=5, column=1, value="New ads since last week");      delta.cell(row=5, column=2, value=len(new_ids))
        delta.cell(row=6, column=1, value="Ads no longer present");        delta.cell(row=6, column=2, value=len(gone_ids))
        delta.cell(row=8, column=1, value="New ad IDs").font = Font(bold=True)
        for i, aid in enumerate(sorted(new_ids), start=9):
            delta.cell(row=i, column=1, value=aid)
    _autosize(delta)

    wb.save(out)
    return out


# ── SEO REPORT ────────────────────────────────────────────────────────────────
def build_seo_report(week_label: Optional[str] = None) -> Path:
    week_label = week_label or current_iso_week()
    out = REPORTS_DIR / f"seo_intel_{week_label}.xlsx"
    data = _load_json(SEO_FILE) or {}

    wb = openpyxl.Workbook()

    # Sheet 1 — Summary (clusters)
    summary = wb.active
    summary.title = "Summary"
    _write_title(summary, f"Listn SEO Intel — {week_label}")
    summary.cell(row=4, column=1, value="Data fetched");  summary.cell(row=4, column=2, value=str(data.get("fetched_date", "")))
    summary.cell(row=6, column=1, value="Content clusters").font = Font(bold=True)
    _write_header(summary, 7, ["Cluster", "Window", "Deadline", "Total volume", "Avg KD", "Keywords"])
    row = 8
    for c in SEO_CLUSTERS:
        kws = c["keywords"]
        total_vol = sum(k[1] for k in kws)
        avg_kd = round(sum(k[2] for k in kws) / len(kws), 1) if kws else 0
        summary.cell(row=row, column=1, value=c["name"])
        summary.cell(row=row, column=2, value=c["window"])
        summary.cell(row=row, column=3, value=str(c["deadline"]) if c["deadline"] else "—")
        summary.cell(row=row, column=4, value=total_vol)
        summary.cell(row=row, column=5, value=avg_kd)
        summary.cell(row=row, column=6, value=len(kws))
        row += 1
    _autosize(summary)

    # Sheet 2 — Raw (DataForSEO competitor keywords)
    raw = wb.create_sheet("Raw")
    competitors = data.get("competitors", {})
    _write_header(raw, 1, ["competitor", "keyword", "search_volume", "position", "keyword_difficulty", "url"])
    row = 2
    for comp, info in competitors.items():
        for kw in info.get("keywords", []):
            raw.cell(row=row, column=1, value=comp)
            raw.cell(row=row, column=2, value=kw.get("keyword"))
            raw.cell(row=row, column=3, value=kw.get("search_volume"))
            raw.cell(row=row, column=4, value=kw.get("position"))
            raw.cell(row=row, column=5, value=kw.get("keyword_difficulty"))
            raw.cell(row=row, column=6, value=kw.get("url"))
            row += 1
    _autosize(raw)

    # Cluster keywords sheet
    cluster_sheet = wb.create_sheet("Cluster Keywords")
    _write_header(cluster_sheet, 1, ["cluster", "window", "keyword", "volume", "kd"])
    row = 2
    for c in SEO_CLUSTERS:
        for kw, vol, kd in c["keywords"]:
            cluster_sheet.cell(row=row, column=1, value=c["name"])
            cluster_sheet.cell(row=row, column=2, value=c["window"])
            cluster_sheet.cell(row=row, column=3, value=kw)
            cluster_sheet.cell(row=row, column=4, value=vol)
            cluster_sheet.cell(row=row, column=5, value=kd)
            row += 1
    _autosize(cluster_sheet)

    # Sheet 3 — Delta
    delta = wb.create_sheet("Delta")
    _write_title(delta, "Week-over-week delta")
    prev = _previous_report("seo_intel_", out)
    if prev is None:
        delta.cell(row=4, column=1, value="No previous week's report found — delta will appear next week.")
    else:
        prev_rows = _read_xlsx_sheet(prev, "Raw")
        prev_kws = {(r[0], r[1]) for r in prev_rows[1:] if r and r[0] and r[1]}
        curr_kws = {(c, k.get("keyword")) for c, info in competitors.items() for k in info.get("keywords", [])}
        new = curr_kws - prev_kws
        gone = prev_kws - curr_kws
        delta.cell(row=4, column=1, value=f"Previous week: {prev.name}")
        delta.cell(row=5, column=1, value="New competitor keywords this week"); delta.cell(row=5, column=2, value=len(new))
        delta.cell(row=6, column=1, value="Keywords no longer ranking");        delta.cell(row=6, column=2, value=len(gone))
    _autosize(delta)

    wb.save(out)
    return out


# ── AI READINESS REPORT ───────────────────────────────────────────────────────
AI_BASELINE = [
    {"name": "Remento",          "llms_txt": "No",  "ai_bots": 0,  "faq_schema": "No",  "article_schema": "No",  "canonical_pct": 8,  "metadesc_pct": 0,  "score": 20},
    {"name": "Heritage Whisper", "llms_txt": "Yes", "ai_bots": 20, "faq_schema": "Yes", "article_schema": "Yes", "canonical_pct": 95, "metadesc_pct": 90, "score": 95},
    {"name": "StoryWorth",       "llms_txt": "No",  "ai_bots": 0,  "faq_schema": "No",  "article_schema": "No",  "canonical_pct": None, "metadesc_pct": None, "score": 15},
    {"name": "Meminto",          "llms_txt": "No",  "ai_bots": 0,  "faq_schema": "No",  "article_schema": "No",  "canonical_pct": None, "metadesc_pct": None, "score": 10},
    {"name": "Listn",            "llms_txt": "No",  "ai_bots": 0,  "faq_schema": "No",  "article_schema": "No",  "canonical_pct": 0,    "metadesc_pct": 0,    "score": 5},
    {"name": "Storykeeper",      "llms_txt": "—",   "ai_bots": "—","faq_schema": "—",   "article_schema": "—",   "canonical_pct": "—",  "metadesc_pct": "—",  "score": "Not yet audited"},
    {"name": "StoriedLife AI",   "llms_txt": "—",   "ai_bots": "—","faq_schema": "—",   "article_schema": "—",   "canonical_pct": "—",  "metadesc_pct": "—",  "score": "Not yet audited"},
    {"name": "LifeEcho",         "llms_txt": "—",   "ai_bots": "—","faq_schema": "—",   "article_schema": "—",   "canonical_pct": "—",  "metadesc_pct": "—",  "score": "Not yet audited"},
    {"name": "Storii",           "llms_txt": "—",   "ai_bots": "—","faq_schema": "—",   "article_schema": "—",   "canonical_pct": "—",  "metadesc_pct": "—",  "score": "Not yet audited"},
]


def _ai_rows() -> list[dict]:
    """Prefer real ai_readiness_latest.json, fall back to baseline."""
    data = _load_json(AI_FILE)
    if data and (data.get("sites") or data.get("results")):
        return data.get("sites") or data.get("results")
    return AI_BASELINE


def build_ai_readiness_report(week_label: Optional[str] = None) -> Path:
    week_label = week_label or current_iso_week()
    out = REPORTS_DIR / f"ai_readiness_{week_label}.xlsx"
    rows = _ai_rows()

    wb = openpyxl.Workbook()

    # Sheet 1 — Summary
    summary = wb.active
    summary.title = "Summary"
    _write_title(summary, f"Listn AI Readiness — {week_label}")
    numeric_scores = [r["score"] for r in rows if isinstance(r.get("score"), (int, float))]
    leader_row = max((r for r in rows if isinstance(r.get("score"), (int, float))), key=lambda r: r["score"], default=None)
    listn_row = next((r for r in rows if r.get("name") == "Listn"), None)
    avg = round(sum(numeric_scores) / len(numeric_scores), 1) if numeric_scores else "—"
    summary.cell(row=4, column=1, value="Listn AI Score");       summary.cell(row=4, column=2, value=(listn_row or {}).get("score", "—"))
    summary.cell(row=5, column=1, value="Industry leader");      summary.cell(row=5, column=2, value=f"{leader_row['name']} ({leader_row['score']})" if leader_row else "—")
    summary.cell(row=6, column=1, value="Industry average");     summary.cell(row=6, column=2, value=avg)
    if isinstance(avg, (int, float)) and listn_row and isinstance(listn_row.get("score"), (int, float)):
        summary.cell(row=7, column=1, value="Listn's gap to leader"); summary.cell(row=7, column=2, value=listn_row["score"] - leader_row["score"])
    _autosize(summary)

    # Sheet 2 — Raw
    raw = wb.create_sheet("Raw")
    _write_header(raw, 1, ["site", "llms_txt", "ai_bots_allowed", "faq_schema", "article_schema",
                           "canonical_pct", "metadesc_pct", "score"])
    for i, r in enumerate(rows, start=2):
        raw.cell(row=i, column=1, value=r.get("name") or r.get("site"))
        raw.cell(row=i, column=2, value=r.get("llms_txt"))
        raw.cell(row=i, column=3, value=r.get("ai_bots"))
        raw.cell(row=i, column=4, value=r.get("faq_schema"))
        raw.cell(row=i, column=5, value=r.get("article_schema"))
        raw.cell(row=i, column=6, value=r.get("canonical_pct"))
        raw.cell(row=i, column=7, value=r.get("metadesc_pct"))
        raw.cell(row=i, column=8, value=r.get("score"))
    _autosize(raw)

    # Sheet 3 — Delta
    delta = wb.create_sheet("Delta")
    _write_title(delta, "Week-over-week delta")
    prev = _previous_report("ai_readiness_", out)
    if prev is None:
        delta.cell(row=4, column=1, value="No previous week's report found — delta will appear next week.")
    else:
        prev_rows_raw = _read_xlsx_sheet(prev, "Raw")
        prev_scores = {row[0]: row[7] for row in prev_rows_raw[1:] if row and row[0]}
        _write_header(delta, 4, ["site", "previous_score", "current_score", "delta"])
        for i, r in enumerate(rows, start=5):
            name = r.get("name") or r.get("site")
            prev_score = prev_scores.get(name)
            curr_score = r.get("score")
            try:
                d = (curr_score - prev_score) if isinstance(curr_score, (int, float)) and isinstance(prev_score, (int, float)) else "—"
            except TypeError:
                d = "—"
            delta.cell(row=i, column=1, value=name)
            delta.cell(row=i, column=2, value=prev_score)
            delta.cell(row=i, column=3, value=curr_score)
            delta.cell(row=i, column=4, value=d)
    _autosize(delta)

    wb.save(out)
    return out


# ── Convenience: build all three ──────────────────────────────────────────────
def build_all_reports(week_label: Optional[str] = None) -> dict[str, Path]:
    return {
        "meta": build_meta_report(week_label),
        "seo":  build_seo_report(week_label),
        "ai":   build_ai_readiness_report(week_label),
    }


def list_reports() -> list[Path]:
    """Return all xlsx files in reports/, newest first."""
    return sorted(REPORTS_DIR.glob("*.xlsx"), reverse=True)


def parse_week_from_filename(path: Path) -> str:
    """Extract '2026-W17' from 'meta_intel_2026-W17.xlsx'."""
    m = re.search(r"(\d{4}-W\d{2})", path.name)
    return m.group(1) if m else path.stem
